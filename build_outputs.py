"""
Case Study Data Analysis — Output Builder
Generates case_study_tables.xlsx and case_study_charts.xlsx
from LC Case Study (March, 2026).pdf data
"""

import math
import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.drawing.image import Image as XLImage
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import matplotlib.ticker as mticker

# ── Constants ──────────────────────────────────────────────────────────────────
CONTRACT_PPM   = 300
PLANNED_HRS    = 303          # hrs/month per machine  (14 hrs/day × ~21.5 days)
METERS_PER_HR  = 2.44         # at 300 PPM, 122.8 EPI
METERS_PER_GARM = 3.45
ASP            = 48.0

# ── Helpers ────────────────────────────────────────────────────────────────────
def bold(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = Font(bold=True)
    return cell

def header_row(ws, row, headers):
    for col, h in enumerate(headers, 1):
        bold(ws, row, col, h)

def autowidth(ws, max_w=55):
    for col in ws.columns:
        w = max(len(str(cell.value or "")) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 2, max_w)

# ══════════════════════════════════════════════════════════════════════════════
# RAW DATA
# ══════════════════════════════════════════════════════════════════════════════

# Section A — Weekly Machine Performance Log
# Cols: Week Ending | Avg PPM | Planned Hrs | Run Hrs | Unplanned DT Hrs |
#       Planned DT Hrs | Garments Started | Garments Completed | Garments Passed QC |
#       Rej Weave Density | Rej Dimensional | Rej Visual Defect | Rej Structural
WEEKLY = [
    ("2025-08-22", 112, 70.0, 31.2, 24.8, 14.0,  68,  52,  28, 8, 6, 5, 5),
    ("2025-08-29", 134, 70.0, 34.5, 20.1, 15.4,  79,  63,  37, 7, 8, 6, 5),
    ("2025-09-05", 148, 70.0, 38.9, 18.6, 12.5,  91,  74,  49, 6, 9, 5, 5),
    ("2025-09-12", 161, 70.0, 41.7, 17.3, 11.0, 104,  88,  61, 8, 7, 7, 5),
    ("2025-09-19", 173, 70.0, 44.2, 14.8, 11.0, 118, 101,  74, 9, 6, 6, 6),
    ("2025-09-26", 189, 70.0, 47.6, 11.9, 10.5, 137, 119,  91, 7, 8, 8, 5),
    ("2025-10-03", 204, 70.0, 50.1, 10.4,  9.5, 153, 136, 108, 6, 9, 7, 6),
    ("2025-10-10", 218, 70.0, 51.8,  9.2,  9.0, 165, 149, 122, 5,10, 6, 6),
    ("2025-10-17", 231, 70.0, 53.4,  8.1,  8.5, 179, 163, 137, 6, 8, 5, 7),
    ("2025-10-24", 244, 70.0, 54.9,  7.6,  7.5, 194, 178, 152, 7, 7, 6, 6),
    ("2025-10-31", 256, 70.0, 56.2,  6.8,  7.0, 208, 192, 167, 5, 8, 5, 7),
    ("2025-11-07", 267, 70.0, 57.3,  5.7,  7.0, 221, 206, 181, 4, 9, 6, 6),
]

DOWNTIME = [
    ("DT-1003-001","2025-10-03","Day","06:42","07:18",36,"Mechanical","Shuttle insertion jam — weft carrier misaligned","Core / Shuttle","Manual re-seat of carrier; recalibrated aperture sensor"),
    ("DT-1003-002","2025-10-03","Day","08:55","09:08",13,"Material","Warp yarn break — end #1847","Creel / Warp Feed","Re-threaded broken end; resumed"),
    ("DT-1003-003","2025-10-03","Day","10:31","10:44",13,"Software","Helix OS diameter control timeout — garment #G-1003-07","Control Software","Restarted diameter control subroutine; lost 2 picks"),
    ("DT-1003-004","2025-10-03","Day","13:02","13:47",45,"Mechanical","Feed-in roller tension variance — overfeed detected","Warp Feed / Rollers","Replaced tension spring on roller #3; recalibrated"),
    ("DT-1003-005","2025-10-03","Day","15:22","15:35",13,"Material","Foreign body in yarn — hard knot from supplier spool #YRN-4412","Creel / Warp Feed","Cut and re-tied; flagged spool for QC review"),
    ("DT-1003-006","2025-10-03","Day","17:08","17:21",13,"Electrical","Wireless power dropout to shuttle — intermittent signal","Core / Power Transfer","Reset wireless power coupling; monitoring"),
    ("DT-1003-007","2025-10-04","Day","07:15","08:03",48,"Mechanical","Aperture ring #4 stuck in narrow position — garment transition fail","Core / Variable Aperture","Disassembled aperture assembly; cleared debris; re-lubricated"),
    ("DT-1003-008","2025-10-04","Day","10:47","11:02",15,"Software","Pattern file load error — style SH-WM-32-BLK","Control Software","Re-uploaded pattern file from server; operator error in file selection"),
    ("DT-1003-009","2025-10-04","Day","14:30","14:52",22,"Mechanical","Haul-off motor speed fluctuation","Haul-off / Finishing","Replaced encoder belt on haul-off motor"),
    ("DT-1003-010","2025-10-05","Day","06:38","07:55",77,"Mechanical","Shuttle insertion mechanism failure — shuttle #SH-003 cracked","Core / Shuttle","Replaced shuttle; full inspection; 2 others showing wear"),
    ("DT-1003-011","2025-10-05","Day","11:20","11:35",15,"Material","Warp yarn break — end #0924 (same spool batch YRN-44xx)","Creel / Warp Feed","Re-threaded; 3rd break from this batch — escalated to procurement"),
    ("DT-1003-012","2025-10-05","Day","14:05","14:18",13,"Software","Diameter control overshoot — waist section too wide by 1.2cm (size 32+)","Control Software","Adjusted PID gain parameter; noted recurring issue on size 32+"),
    ("DT-1003-013","2025-10-06","Day","08:22","08:48",26,"Mechanical","Creel yarn tension uneven — multiple ends slack","Creel / Warp Feed","Re-tensioned creel positions 1800–1850; found broken spring"),
    ("DT-1003-014","2025-10-06","Day","12:44","13:20",36,"Changeover","Style changeover SH-WM-32-BLK → SH-WM-34-NVY","All","Pattern load, aperture recalibration, weft yarn swap"),
    ("DT-1003-015","2025-10-06","Day","16:30","16:48",18,"Mechanical","Abnormal vibration detected — core housing","Core / Structure","Tightened core mounting bolts; vibration within spec after"),
    ("DT-1003-016","2025-10-07","Day","07:02","07:25",23,"Electrical","Power supply unit fault — PSU #2 thermal shutdown","Electrical / PSU","Waited for cool-down; restarted; PSU fan partially blocked"),
    ("DT-1003-017","2025-10-07","Day","09:48","10:03",15,"Material","Weft yarn splice failure — loose join from supplier","Weft Feed","Re-spliced; flagged to supplier"),
    ("DT-1003-018","2025-10-07","Day","13:15","13:33",18,"Software","MOS data logging failure — machine state not recording","Control Software / MOS","Restarted logging daemon; ~15 min of data lost"),
    ("DT-1003-019","2025-10-07","Day","15:55","17:10",75,"Mechanical","Core aperture ring #2 fractured — metal fatigue","Core / Variable Aperture","Emergency replacement; full aperture assembly inspection required"),
]

QC_DATA = [
    ("G-1003-01","SH-WM-32-BLK",32,"2025-10-03",121.4,+0.3,+0.2,-0.1,+0.2,1,"PASS","PASS","—"),
    ("G-1003-02","SH-WM-32-BLK",32,"2025-10-03",119.7,+0.1,+0.4,-0.2,+0.1,0,"PASS","PASS","—"),
    ("G-1003-03","SH-WM-32-BLK",32,"2025-10-03",115.2,+0.2,+0.3,+0.1,+0.3,0,"PASS","FAIL","Weave density below min (118 EPI)"),
    ("G-1003-04","SH-WM-32-BLK",32,"2025-10-03",123.1,+0.8,+0.1,-0.1,+0.2,2,"PASS","FAIL","Waist dim out of tolerance (+0.8 cm)"),
    ("G-1003-05","SH-WM-34-NVY",34,"2025-10-03",122.5,-0.2,+0.3,+0.4,+0.1,0,"PASS","PASS","—"),
    ("G-1003-06","SH-WM-34-NVY",34,"2025-10-03",124.8,+0.1,-0.6,-0.3,-0.1,1,"PASS","FAIL","Hip dim out of tolerance (-0.6 cm)"),
    ("G-1003-07","SH-WM-32-BLK",32,"2025-10-03",120.3,+0.2,+0.1,-0.2,+0.3,4,"PASS","FAIL","Visual defect points exceed limit"),
    ("G-1003-08","SH-WM-34-NVY",34,"2025-10-04",122.1,+0.1,+0.2,+0.1,+0.2,0,"PASS","PASS","—"),
    ("G-1003-09","SH-WM-32-BLK",32,"2025-10-04",123.6,-0.3,+0.4,-0.1,+0.1,1,"PASS","PASS","—"),
    ("G-1003-10","SH-WM-32-BLK",32,"2025-10-04",117.1,+0.2,+0.1,+0.3,-0.2,0,"PASS","FAIL","Weave density below min (118 EPI)"),
    ("G-1003-11","SH-WM-34-NVY",34,"2025-10-04",121.9,+0.4,+0.3,-0.4,+0.5,0,"PASS","PASS","—"),
    ("G-1003-12","SH-WM-34-NVY",34,"2025-10-04",122.4,+0.1,+0.2,-0.1,+0.3,0,"FAIL","FAIL","Warp float >3mm in left leg panel"),
    ("G-1003-13","SH-WM-32-BLK",32,"2025-10-04",123.0,+0.2,+0.1,+0.1,+0.2,1,"PASS","PASS","—"),
    ("G-1003-14","SH-WM-36-BLK",36,"2025-10-05",121.5,+0.1,-0.3,+0.2,+0.1,0,"PASS","PASS","—"),
    ("G-1003-15","SH-WM-36-BLK",36,"2025-10-05",118.4,+0.7,+0.9,+0.6,+0.8,2,"PASS","FAIL","Multiple dims out of tolerance"),
    ("G-1003-16","SH-WM-32-BLK",32,"2025-10-05",123.3,+0.3,+0.2,-0.2,+0.1,0,"PASS","PASS","—"),
    ("G-1003-17","SH-WM-34-NVY",34,"2025-10-05",120.8,-0.1,+0.3,+0.1,+0.2,3,"PASS","FAIL","Visual defect points exceed limit"),
    ("G-1003-18","SH-WM-32-BLK",32,"2025-10-06",122.6,+0.2,+0.1,-0.1,+0.3,0,"PASS","PASS","—"),
    ("G-1003-19","SH-WM-34-NVY",34,"2025-10-06",123.8,+0.3,+0.4,-0.3,+0.2,1,"PASS","PASS","—"),
    ("G-1003-20","SH-WM-36-BLK",36,"2025-10-06",114.9,+1.1,+0.8,+0.4,+0.7,1,"PASS","FAIL","Weave density below min + waist out of tolerance"),
    ("G-1003-21","SH-WM-32-BLK",32,"2025-10-06",122.2,+0.1,+0.2,+0.2,+0.1,0,"PASS","PASS","—"),
    ("G-1003-22","SH-WM-34-NVY",34,"2025-10-07",121.1,-0.2,+0.1,+0.3,+0.2,0,"PASS","PASS","—"),
    ("G-1003-23","SH-WM-32-BLK",32,"2025-10-07",122.9,+0.4,+0.3,-0.2,+0.4,2,"PASS","PASS","—"),
    ("G-1003-24","SH-WM-36-BLK",36,"2025-10-07",119.2,-0.4,+0.5,+0.3,-0.3,0,"PASS","PASS","—"),
    ("G-1003-25","SH-WM-34-NVY",34,"2025-10-07",122.0,+0.1,+0.2,+0.1,+0.1,0,"PASS","PASS","—"),
    ("G-1003-26","SH-WM-32-BLK",32,"2025-10-07",123.5,+0.3,+0.1,-0.1,+0.2,1,"PASS","PASS","—"),
]

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK 1 — TABLES
# ══════════════════════════════════════════════════════════════════════════════
wb = openpyxl.Workbook()
wb.remove(wb.active)

# ── Sheet 1: Weekly Performance Log ──────────────────────────────────────────
ws = wb.create_sheet("Weekly Performance Log")
bold(ws, 1, 1, "Section A — Weekly Machine Performance Log (Helix Unit 1)")
bold(ws, 2, 1, "Machine configured: 14-hr production day, 5 days/week. Contract target speed: 300 PPM.")
hdrs = ["Week Ending","Avg PPM","Planned Hrs","Run Hrs","Unplanned DT Hrs","Planned DT Hrs",
        "Garments Started","Garments Completed","Garments Passed QC",
        "Reject - Weave Density","Reject - Dimensional","Reject - Visual Defect","Reject - Structural",
        "Availability %","Performance %","Quality Yield %","OEE % (calc)"]
header_row(ws, 4, hdrs)

for r, row in enumerate(WEEKLY, 5):
    for c, v in enumerate(row, 1):
        ws.cell(row=r, column=c, value=v)
    avail = row[3] / row[2]
    perf  = row[1] / CONTRACT_PPM
    qual  = row[8] / row[7] if row[7] > 0 else 0
    oee   = avail * perf * qual
    ws.cell(row=r, column=14, value=round(avail*100, 1))
    ws.cell(row=r, column=15, value=round(perf*100, 1))
    ws.cell(row=r, column=16, value=round(qual*100, 1))
    ws.cell(row=r, column=17, value=round(oee*100, 1))

bold(ws, 18, 1, "Note: OEE calculated as Availability (Run Hrs / Planned Hrs) × Performance (Avg PPM / 300) × Quality Yield (Passed QC / Completed)")
autowidth(ws)

# ── Sheet 2: Downtime Events ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Downtime Events")
bold(ws2, 1, 1, "Section B — Downtime Event Log (Helix Unit 1 — Sample Week: 2025-10-03)")
header_row(ws2, 3, ["Event ID","Date","Shift","Start Time","End Time","Duration (min)",
                     "Stop Category","Stop Reason","Subsystem","Resolution"])
for r, row in enumerate(DOWNTIME, 4):
    for c, v in enumerate(row, 1):
        ws2.cell(row=r, column=c, value=v)

sr = len(DOWNTIME) + 6
bold(ws2, sr,   1, "Week Summary (2025-10-03):")
ws2.cell(row=sr+1, column=1, value="Total unplanned downtime events:")
ws2.cell(row=sr+1, column=2, value=19)
ws2.cell(row=sr+2, column=1, value="Total unplanned downtime (min):")
ws2.cell(row=sr+2, column=2, value=624)
ws2.cell(row=sr+3, column=1, value="Total unplanned downtime (hrs):")
ws2.cell(row=sr+3, column=2, value=10.4)
ws2.cell(row=sr+4, column=1, value="Top category: Mechanical (10 events, 396 min)")
ws2.cell(row=sr+5, column=1, value="Top subsystem: Core / Shuttle + Core / Variable Aperture (5 events, 282 min)")
ws2.cell(row=sr+6, column=1, value="Material-related stops: 4 events (54 min)")
ws2.cell(row=sr+7, column=1, value="Software-related stops: 4 events (61 min)")
autowidth(ws2)

# ── Sheet 3: Garment QC Data ──────────────────────────────────────────────────
ws3 = wb.create_sheet("Garment QC Data")
bold(ws3, 1, 1, "Section C — Quality Inspection Data (Garment-Level, Week of 2025-10-03)")
bold(ws3, 2, 1, "Customer A Acceptance Criteria: EPI 118–128 (target 122.8) | Dimensional ±0.5 cm | Visual ≤2 minor pts | Structural: seam pull ≥22 lbs/inch")
header_row(ws3, 4, ["Garment ID","Style","Size","Date Completed","EPI (Measured)",
                     "Waist (cm vs spec)","Hip (cm vs spec)","Inseam (cm vs spec)","Outseam (cm vs spec)",
                     "Visual Defect Points","Structural Pass?","Overall QC Result","Reject Reason"])
for r, row in enumerate(QC_DATA, 5):
    for c, v in enumerate(row, 1):
        ws3.cell(row=r, column=c, value=v)

sr3 = len(QC_DATA) + 7
bold(ws3, sr3,   1, "Week Summary:")
ws3.cell(row=sr3+1, column=1, value="Total garments completed:"); ws3.cell(row=sr3+1, column=2, value=26)
ws3.cell(row=sr3+2, column=1, value="Passed QC:");                ws3.cell(row=sr3+2, column=2, value=18)
ws3.cell(row=sr3+3, column=1, value="QC Yield %:");               ws3.cell(row=sr3+3, column=2, value="69.2%")
ws3.cell(row=sr3+4, column=1, value="Reject breakdown: 3 weave density, 3 dimensional, 2 visual defects")
autowidth(ws3)

# ── Sheet 4: Unit Economics (BOM) ─────────────────────────────────────────────
ws4 = wb.create_sheet("Unit Economics")
bold(ws4, 1, 1, "Section D — Bill of Materials: Per-Garment COGS (Style SH-WM-32-BLK)")
bold(ws4, 2, 1, "At current volume (~150 garments/month on one machine)")
header_row(ws4, 4, ["Cost Category","Line Item","Unit","Qty/Garment","Unit Cost","Cost/Garment ($)","Notes"])

BOM = [
    ("Raw Material","Warp yarn (cotton, 16/1 Ne)","kg",0.42,"$8.50/kg",3.57,"Domestic supplier; dyed-to-order"),
    ("","Weft yarn (cotton/elastane 97/3)","kg",0.18,"$12.30/kg",2.21,"Specialty blend; single source"),
    ("","Waistband elastic + hardware","set",1,"$0.85",0.85,""),
    ("","Button + rivets","set",1,"$0.42",0.42,""),
    ("","Label + hangtag","set",1,"$0.18",0.18,"Customer-supplied artwork"),
    ("","Thread (finishing/hemming)","m",12,"$0.003/m",0.04,""),
    ("","Packaging (poly bag + carton)","set",1,"$0.32",0.32,""),
    ("Raw Material Subtotal","","","","",7.59,""),
    ("Direct Labor","Machine operator (weaving)","hrs",0.19,"$28.00/hr",5.32,"Shared across 3 garments on loom simultaneously"),
    ("","Finishing (off-loom sewing)","hrs",0.25,"$24.00/hr",6.00,"Waistband attach, hem, pocket set, hardware"),
    ("","QC inspection","hrs",0.08,"$26.00/hr",2.08,"Per-garment inspection + measurement"),
    ("Direct Labor Subtotal","","","","",13.40,""),
    ("Machine Overhead","Machine depreciation","garment",1,"—",4.80,"Straight-line over 7-yr life; per-unit at current throughput"),
    ("","Electricity","kWh",3.2,"$0.14/kWh",0.45,""),
    ("","Compressed air","—","—","—",0.12,""),
    ("","Consumables (shuttles, aperture rings, sensors)","garment",1,"—",2.35,"Shuttle replacement cycle ~200 garments"),
    ("","Maintenance labor (allocated)","garment",1,"—",1.85,"~0.5 FTE dedicated to machine maintenance"),
    ("Machine Overhead Subtotal","","","","",9.57,""),
    ("Facility","Rent (allocated)","garment",1,"—",1.20,"Oakland facility; machine footprint + staging"),
    ("","Insurance + utilities (allocated)","garment",1,"—",0.55,""),
    ("Facility Subtotal","","","","",1.75,""),
    ("Scrap & Rework","Yarn waste (est. 3.2% of material)","—","—","—",0.24,"Near-net-shape; minimal vs. traditional 14%"),
    ("","Rework cost (garments requiring re-finishing)","—","—","—",0.95,"~8% of garments require minor rework post-QC"),
    ("Scrap/Rework Subtotal","","","","",1.19,""),
    ("TOTAL COGS PER GARMENT","","","","",33.50,"At current volume (~150/month)"),
]
for r, row in enumerate(BOM, 5):
    for c, v in enumerate(row, 1):
        cell = ws4.cell(row=r, column=c, value=v)
        if "Subtotal" in str(row[0]) or row[0] == "TOTAL COGS PER GARMENT":
            cell.font = Font(bold=True)
autowidth(ws4)

# ── Sheet 5: Scale Sensitivity ────────────────────────────────────────────────
ws5 = wb.create_sheet("Scale Sensitivity")
bold(ws5, 1, 1, "Section E — Unit Economics: Scale Sensitivity Analysis (Company Projections)")
bold(ws5, 2, 1, "ASP: $48 blended (Customer A mass-market + pipeline premium brand customers)")
header_row(ws5, 4, ["Scenario","Machines","Garments/Month","Yarn Cost ($)","Labor Cost ($)",
                     "Machine Overhead ($)","Facility ($)","Scrap/Rework ($)","Total COGS ($)","Gross Margin (at $48 ASP)"])
SCALE = [
    ("Current (actual)",            1,   "~150",  7.59, 13.40, 9.57, 1.75, 1.19, 33.50, "30.2%"),
    ("Near-term target (Q2 2026)",  1,   "~600",  7.20,  9.80, 4.10, 0.85, 0.75, 22.70, "52.7%"),
    ("Mid-term (Q4 2026)",          3,  "~3,000", 6.40,  7.20, 2.80, 0.45, 0.50, 17.35, "63.9%"),
    ("At scale (2028)",           "10+","~15,000", 5.80,  4.50, 1.60, 0.25, 0.30, 12.45, "74.1%"),
]
for r, row in enumerate(SCALE, 5):
    for c, v in enumerate(row, 1):
        ws5.cell(row=r, column=c, value=v)

bold(ws5, 11, 1, "Key Assumptions in Company Model:")
ASSM = [
    "• Labor reduction: finishing automation (robotic waistband attachment), ML-based visual QC, 1 operator managing 3 machines at scale",
    "• Machine overhead: shuttle life ~200 → ~2,000 garments (10x); aperture ring life ~500 → ~5,000 garments (10x); depreciation spread over higher volume",
    "• Yarn cost reduction: volume procurement discounts and second-source qualification for weft yarn",
    "• ASP of $48 is blended; Customer A (mass retailer) is lower ASP, pipeline premium brands are higher",
]
for r, note in enumerate(ASSM, 12):
    ws5.cell(row=r, column=1, value=note)
    ws5.merge_cells(f"A{r}:J{r}")
autowidth(ws5)

# ── Sheet 6: Customer A Contract ──────────────────────────────────────────────
ws6 = wb.create_sheet("Customer A Contract")
bold(ws6, 1, 1, "Section F — Customer A: Contractual Production Requirements (Pilot Production Agreement)")
header_row(ws6, 3, ["Parameter", "Requirement"])
CONTRACT = [
    ("Pilot period",                      "6 months (Jan 2026 — Jun 2026)"),
    ("Minimum monthly delivery",          "1,250 garments"),
    ("Product",                           "Men's woven shorts, 4 styles × 5 sizes"),
    ("On-loom fabric per garment",        "3.45 meters (average across sizes)"),
    ("Required machine speed",            "≥ 300 PPM (contract minimum)"),
    ("Required OEE",                      "≥ 50% (availability 78% × performance 80% × quality 80%)"),
    ("Quality standard",                  "AQL 2.5 major / AQL 4.0 minor; 0 critical"),
    ("Delivery schedule",                 "Weekly shipments; ±5% tolerance on weekly volume"),
    ("Penalty clause",                    "If 3 consecutive weeks below 80% of target volume, customer may terminate pilot"),
    ("Exclusivity",                       "Non-exclusive; company retains right to produce for other customers on other machines"),
    ("Pricing",                           "Fixed price per garment for pilot period (not disclosed)"),
    ("Extension option",                  "If pilot targets met, 12-month production agreement at 5,000 garments/month"),
]
for r, row in enumerate(CONTRACT, 4):
    ws6.cell(row=r, column=1, value=row[0]).font = Font(bold=True)
    ws6.cell(row=r, column=2, value=row[1])

bold(ws6, 18, 1, "Implied Production Math (at Contract Terms):")
MATH = [
    "1,250 garments × 3.45 m = 4,312.5 meters of on-loom fabric/month",
    "At 300 PPM and 122.8 EPI on-loom density: ~2.44 meters/hour production rate",
    "4,312.5 m ÷ 2.44 m/hr = 1,767 production hours required/month",
    "Available hours: 14 hrs/day × ~21.5 working days = 303 hrs/month per machine",
    "At 50% OEE: effective production ≈ 151.5 hrs → ~370 meters → ~107 garments/month per machine",
    "Gap: 1 machine at 50% OEE produces ~107 garments/month vs. 1,250 required (see Production Gap Analysis sheet)",
]
for r, note in enumerate(MATH, 19):
    ws6.cell(row=r, column=1, value=note)
    ws6.merge_cells(f"A{r}:B{r}")
autowidth(ws6)

# ── Sheet 7: Machine Build Pipeline ──────────────────────────────────────────
ws7 = wb.create_sheet("Machine Build Pipeline")
bold(ws7, 1, 1, "Section G — Machine Build & Deployment Pipeline")
header_row(ws7, 3, ["Machine Unit","Status","Location","Assigned Customer","Est. Commissioning","Notes"])
PIPELINE = [
    ("Helix-001","Operational",      "HQ facility","Internal R&D / Customer A pilot","Commissioned 2025-06","Primary production unit; all performance data above"),
    ("Helix-002","Assembly (85%)",   "HQ facility","Customer A pilot",               "2026-01 (target)",    "Delayed 6 weeks from original schedule; waiting on custom aperture ring set"),
    ("Helix-003","Assembly (40%)",   "HQ facility","Customer A pilot",               "2026-03 (target)",    "Long-lead components ordered"),
    ("Helix-004","Procurement",      "—",          "Customer B (premium outdoor brand)","2026-Q2 (target)", "Deposit received; build not started"),
    ("Helix-005","Design",           "—",          "Unallocated",                    "2026-Q3 (target)",    "Gen 2 design incorporating learnings from 001–003"),
]
for r, row in enumerate(PIPELINE, 4):
    for c, v in enumerate(row, 1):
        ws7.cell(row=r, column=c, value=v)
autowidth(ws7)

# ── Sheet 8: Engineering Change Log ──────────────────────────────────────────
ws8 = wb.create_sheet("Engineering Change Log")
bold(ws8, 1, 1, "Section H — Engineering Change Log (Helix-001, Last 90 Days)")
header_row(ws8, 3, ["ECO #","Date Filed","Subsystem","Description","Status","Impact on Production"])
ECO = [
    ("ECO-2025-041","2025-08-14","Core / Shuttle",           "Redesigned shuttle insertion ramp angle from 12° to 8° to reduce jamming",         "Implemented 2025-09-02","Shuttle jams reduced ~40%; PPM improved"),
    ("ECO-2025-042","2025-08-22","Core / Variable Aperture",  "Added hardened steel wear surface to aperture ring contact points",                "Implemented 2025-09-10","Ring replacement interval up from ~300 to ~500 garments"),
    ("ECO-2025-043","2025-09-05","Warp Feed",                 "Redesigned creel tension mechanism for positions 1800–2592 (outer ring)",          "Implemented 2025-09-18","Reduced warp break rate by ~25% in outer positions"),
    ("ECO-2025-044","2025-09-15","Control Software",          "Updated PID tuning for diameter control — reduced overshoot on size 34+",           "Implemented 2025-09-22","Dimensional reject rate on size 34+ down ~30%"),
    ("ECO-2025-045","2025-09-28","Core / Shuttle",            "New shuttle material (ceramic composite vs. hardened steel) — prototype testing",   "In Testing",            "If successful, est. shuttle life 3–5× improvement"),
    ("ECO-2025-046","2025-10-05","Core / Variable Aperture",  "Root cause analysis of aperture ring #2 fracture (fatigue); proposed material change","In Review",           "Single-source for aperture rings; 14-week lead time"),
    ("ECO-2025-047","2025-10-12","Haul-off / Finishing",      "New encoder belt material for haul-off motor (Kevlar-reinforced vs. rubber)",       "Implemented 2025-10-20","Belt replacement interval from ~30 days to est. ~90 days"),
    ("ECO-2025-048","2025-10-18","Control Software",          "ML-based visual inspection prototype — camera integration for off-loom QC",         "In Development",        "Target: reduce QC inspection time from 5 min to <2 min per garment"),
    ("ECO-2025-049","2025-10-25","Core / Power Transfer",     "Redesigned wireless power antenna for shuttle — increase coupling efficiency 72%→85%","In Design",           "Would reduce wireless dropouts; requires shuttle redesign"),
    ("ECO-2025-050","2025-11-01","All",                       "Gen 2 machine architecture study — modular core for easier aperture ring servicing", "In Design",             "Not applicable to machines 001–003; targeted for machine 005+"),
]
for r, row in enumerate(ECO, 4):
    for c, v in enumerate(row, 1):
        ws8.cell(row=r, column=c, value=v)
autowidth(ws8)

# ── Sheet 9: Forward Milestones ───────────────────────────────────────────────
ws9 = wb.create_sheet("Forward Milestones")
bold(ws9, 1, 1, "Section 3.5 — Forward Milestone Targets (Machine Performance Roadmap)")
header_row(ws9, 3, ["Milestone","Speed","Availability","Performance","Quality","OEE","Meters/Month"])
MILESTONES = [
    ("Q4 '25 (achieved)", "150 PPM", "75%", "90%", "75%", "~51%", "1,120"),
    ("Q2 '26 (target)",   "300 PPM", "75%", "90%", "75%", "~51%", "2,220"),
    ("Q3 '26 (target)",   "400 PPM", "85%", "95%", "90%", "~73%", "4,220"),
]
for r, row in enumerate(MILESTONES, 4):
    for c, v in enumerate(row, 1):
        ws9.cell(row=r, column=c, value=v)

bold(ws9, 9, 1, "Technology Readiness Level (TRL) Assessment:")
header_row(ws9, 10, ["Milestone", "TRL Assessment"])
TRL = [
    ("At Series B",          "TRL-4: Component/system validation in laboratory environment"),
    ("Today",                "TRL-5: System validation in relevant (pilot) environment"),
    ("Target by next raise", "TRL-7: Full-scale system demonstrated in relevant environment"),
]
for r, row in enumerate(TRL, 11):
    for c, v in enumerate(row, 1):
        ws9.cell(row=r, column=c, value=v)

bold(ws9, 16, 1, "Notes:")
NOTES9 = [
    "Q4 2025 milestone (150 PPM, ~51% OEE) confirmed achieved October 30th",
    "Q2 2026 requires sustaining 50% OEE at 300 PPM over long-duration runs in external customer factory",
    "Q3 2026 represents full production capacity of the long-term machine design",
]
for r, note in enumerate(NOTES9, 17):
    ws9.cell(row=r, column=1, value=note)
autowidth(ws9)

# ── Sheet 10: LCA Comparison ──────────────────────────────────────────────────
ws10 = wb.create_sheet("LCA Comparison")
bold(ws10, 1, 1, "Section J — Life Cycle Assessment: Helix vs. Traditional Manufacturing")
bold(ws10, 2, 1, "Third-party verified. Cradle-to-gate boundary (yarn production → finished garment at distribution centre).")
header_row(ws10, 4, ["Parameter","Helix System","Traditional (cut-and-sew, SE Asia)"])
LCA = [
    ("Fabric waste (% of input material)",               "3.2%",                         "14.0%"),
    ("Sell-through rate (assumed)",                      "100%",                          "70%"),
    ("Manufacturing steps",                              "1 (weave) + 1 (finishing)",     "5+ (beam warp → weave → finish → cut → sew)"),
    ("Transport distance (factory to DC)",               "~50 km (nearshore)",            "~15,000 km (transoceanic)"),
    ("Energy per garment (MJ)",                          18.4,                            36.1),
    ("Water per garment (litres)",                       32,                              52),
    ("CO₂e per garment (kg)",                            3.8,                             8.1),
    ("Energy reduction (Helix vs. traditional)",         "-49%",                          "—"),
    ("Water reduction (Helix vs. traditional)",          "-39%",                          "—"),
    ("CO₂e reduction (Helix vs. traditional)",           "-53%",                          "—"),
]
for r, row in enumerate(LCA, 5):
    for c, v in enumerate(row, 1):
        ws10.cell(row=r, column=c, value=v)

bold(ws10, 17, 1, "LCA Notes:")
NOTES10 = [
    "• Does not include consumer use phase or end-of-life",
    "• Traditional system assumes global industry average; top-performing factories would show smaller gap",
    "• Helix 100% sell-through assumes on-demand production; actual depends on customer ordering patterns",
    "• 70% traditional sell-through is industry average; premium brands ~80–85%, fast fashion ~60–65%",
]
for r, note in enumerate(NOTES10, 18):
    ws10.cell(row=r, column=1, value=note)
    ws10.merge_cells(f"A{r}:C{r}")
autowidth(ws10)

# ── Sheet 11: Production Gap Analysis ────────────────────────────────────────
ws11 = wb.create_sheet("Production Gap Analysis")
bold(ws11, 1, 1, "Task 2 — Production Gap Analysis: Machines Required to Meet Customer A Contract Targets")
bold(ws11, 2, 1, f"Inputs: {PLANNED_HRS} hrs/month per machine | {CONTRACT_PPM} PPM | {METERS_PER_HR} m/hr production rate | {METERS_PER_GARM} m/garment")

gap_hdrs = ["OEE Level","Effective Hrs / Machine / Month","Meters / Machine / Month",
            "Garments / Machine / Month","Machines Required (rounded up)","Notes"]

bold(ws11, 4, 1, "(a)  Pilot Target: 1,250 garments / month")
header_row(ws11, 5, gap_hdrs)
for i, oee_pct in enumerate([40, 50, 60, 70]):
    oee = oee_pct / 100
    eff  = PLANNED_HRS * oee
    mts  = eff * METERS_PER_HR
    garm = mts / METERS_PER_GARM
    mach = math.ceil(1250 / garm)
    note = "Contract minimum OEE" if oee_pct == 50 else ""
    r = 6 + i
    vals = [f"{oee_pct}%", round(eff,1), round(mts,1), round(garm,1), mach, note]
    for c, v in enumerate(vals, 1):
        cell = ws11.cell(row=r, column=c, value=v)
        if oee_pct == 50:
            cell.font = Font(bold=True)

bold(ws11, 12, 1, "(b)  Extension Target: 5,000 garments / month")
header_row(ws11, 13, gap_hdrs)
for i, oee_pct in enumerate([40, 50, 60, 70]):
    oee = oee_pct / 100
    eff  = PLANNED_HRS * oee
    mts  = eff * METERS_PER_HR
    garm = mts / METERS_PER_GARM
    mach = math.ceil(5000 / garm)
    note = "Contract minimum OEE" if oee_pct == 50 else ""
    r = 14 + i
    vals = [f"{oee_pct}%", round(eff,1), round(mts,1), round(garm,1), mach, note]
    for c, v in enumerate(vals, 1):
        cell = ws11.cell(row=r, column=c, value=v)
        if oee_pct == 50:
            cell.font = Font(bold=True)

bold(ws11, 20, 1, "Key Observations:")
OBS = [
    "• At contract minimum OEE (50%): 12 machines needed for pilot (1,250/month), 47 machines for extension (5,000/month)",
    "• Company currently has 3 machines in pipeline (Helix-001 operational; -002 and -003 in assembly) — far short of 12 at 50% OEE",
    "• Latest data (week of 2025-11-07) shows Helix-001 OEE at ~70% (calculated) — if sustained, pilot would require ~9 machines",
    "• Bridging the gap requires either: higher sustained OEE than contract minimum, additional machine builds, or revised pilot scope",
    "• Extension target (5,000/month) implies 34–47 machines depending on OEE; major scale-up of machine production required",
]
for r, note in enumerate(OBS, 21):
    ws11.cell(row=r, column=1, value=note)
    ws11.merge_cells(f"A{r}:F{r}")
autowidth(ws11, max_w=35)

# ── Sheet 12: COGS Sensitivity ────────────────────────────────────────────────
ws12 = wb.create_sheet("COGS Sensitivity")
bold(ws12, 1, 1, "Task 4 — COGS Sensitivity: Impact of Partial Scale Assumptions on Gross Margin")
bold(ws12, 2, 1, "Tests what happens if company cost improvement assumptions are only partially achieved vs. company base case")

# Component cost assumptions
SHUTTLE_COST  = 1.175  # $/garment at 200-garment shuttle life (50% of $2.35 consumables)
APERTURE_COST = 0.823  # $/garment at 500-garment aperture ring life (35% of $2.35)
SENSOR_COST   = 0.353  # $/garment fixed (15% of $2.35)

# Q2 2026 base case: Total COGS $22.70
# Non-consumable machine overhead (depreciation, electricity, compressed air, maintenance) at Q2 level
# = Company machine OH ($4.10) − consumables at 10x improvement
CONS_BASE_Q2  = SHUTTLE_COST/10 + APERTURE_COST/10 + SENSOR_COST   # base case 10x for both
NON_CONS_OH_Q2 = 4.10 - CONS_BASE_Q2

CONS_BASE_Q4  = SHUTTLE_COST/10 + APERTURE_COST/10 + SENSOR_COST
NON_CONS_OH_Q4 = 2.80 - CONS_BASE_Q4

# Labor
LABOR_Q2_AUTO    = 9.80   # company base with full finishing automation
LABOR_Q2_DELAYED = 11.80  # finishing stays manual; some benefit from volume scaling only
LABOR_Q4_AUTO    = 7.20   # company base with full automation
LABOR_Q4_DELAYED = 10.20  # finishing stays manual; partial benefit from 3-machine operator scaling

# Fixed inputs
YARN_Q2     = 7.20;  FAC_Q2  = 0.85;  SCRAP_Q2  = 0.75
YARN_Q4     = 6.40;  FAC_Q4  = 0.45;  SCRAP_Q4  = 0.50

bold(ws12, 4, 1, "Assumption: Consumables ($2.35/garment at current) split — Shuttles 50% ($1.175), Aperture Rings 35% ($0.823), Sensors/Other 15% ($0.353)")
bold(ws12, 5, 1, "Company base case: Shuttle life 200→2,000 garments (10x); Aperture ring life 500→5,000 garments (10x)")
ws12.merge_cells("A4:N4")
ws12.merge_cells("A5:N5")

SENS_HDRS = [
    "Scenario","Scenario Details",
    "Q2 2026\nConsumables ($/garm)","Q2 2026\nMachine OH ($/garm)","Q2 2026\nLabor ($/garm)",
    "Q2 2026\nTotal COGS ($)","Q2 2026\nGross Margin","Q2 2026\nvs. Base",
    "Q4 2026\nConsumables ($/garm)","Q4 2026\nMachine OH ($/garm)","Q4 2026\nLabor ($/garm)",
    "Q4 2026\nTotal COGS ($)","Q4 2026\nGross Margin","Q4 2026\nvs. Base",
]
header_row(ws12, 7, SENS_HDRS)

SCENARIOS = [
    ("Base Case (Company)",                     10, 10, False),
    ("Sensitivity A: Shuttle 3x only",           3, 10, False),
    ("Sensitivity B: Aperture ring 2x only",    10,  2, False),
    ("Sensitivity C: Both partial (3x + 2x)",    3,  2, False),
    ("Sensitivity D: Both partial + labor delayed", 3, 2, True),
]

SCENARIO_DETAIL = {
    "Base Case (Company)":                     "Shuttle 10x (→2,000 garm), Aperture 10x (→5,000 garm), Full labor automation",
    "Sensitivity A: Shuttle 3x only":          "Shuttle 3x (→600 garm), Aperture 10x, Full labor automation",
    "Sensitivity B: Aperture ring 2x only":    "Shuttle 10x, Aperture 2x (→1,000 garm), Full labor automation",
    "Sensitivity C: Both partial (3x + 2x)":   "Shuttle 3x (→600 garm), Aperture 2x (→1,000 garm), Full labor automation",
    "Sensitivity D: Both partial + labor delayed": "Shuttle 3x (→600 garm), Aperture 2x (→1,000 garm), Finishing automation NOT implemented",
}

# Base case gross margins for delta calculation
base_q2_gm = (ASP - 22.70) / ASP
base_q4_gm = (ASP - 17.35) / ASP

for i, (name, sh_mult, ap_mult, labor_delay) in enumerate(SCENARIOS):
    cons = SHUTTLE_COST/sh_mult + APERTURE_COST/ap_mult + SENSOR_COST

    moh_q2 = NON_CONS_OH_Q2 + cons
    moh_q4 = NON_CONS_OH_Q4 + cons
    lab_q2 = LABOR_Q2_DELAYED if labor_delay else LABOR_Q2_AUTO
    lab_q4 = LABOR_Q4_DELAYED if labor_delay else LABOR_Q4_AUTO

    tot_q2 = YARN_Q2 + lab_q2 + moh_q2 + FAC_Q2 + SCRAP_Q2
    tot_q4 = YARN_Q4 + lab_q4 + moh_q4 + FAC_Q4 + SCRAP_Q4
    gm_q2  = (ASP - tot_q2) / ASP
    gm_q4  = (ASP - tot_q4) / ASP
    dq2 = gm_q2 - base_q2_gm
    dq4 = gm_q4 - base_q4_gm

    r = 8 + i
    vals = [
        name, SCENARIO_DETAIL[name],
        round(cons,2), round(moh_q2,2), round(lab_q2,2), round(tot_q2,2),
        f"{gm_q2:.1%}", "—" if i==0 else f"{dq2:+.1%}",
        round(cons,2), round(moh_q4,2), round(lab_q4,2), round(tot_q4,2),
        f"{gm_q4:.1%}", "—" if i==0 else f"{dq4:+.1%}",
    ]
    for c, v in enumerate(vals, 1):
        cell = ws12.cell(row=r, column=c, value=v)
        if i == 0:
            cell.font = Font(bold=True)

bold(ws12, 15, 1, "Notes:")
NOTES12 = [
    "• Consumables split is estimated from BOM data: Shuttles ~50%, Aperture Rings ~35%, Sensors/Other ~15% of the $2.35 total",
    "• Non-consumable machine overhead (depreciation, electricity, compressed air, maintenance labor) held at company base case levels",
    "• 'Labor Delayed' assumes finishing automation not implemented: finishing labor (~$6.00/garment) stays at current rate; machine operator shows some volume benefit",
    "• Yarn, facility, and scrap/rework costs held at company base case projections across all sensitivity scenarios",
    "• Gross margin calculated at $48 blended ASP. All figures are per-garment COGS.",
    f"• Company base case gross margins: Q2 2026 = {base_q2_gm:.1%}, Q4 2026 = {base_q4_gm:.1%}",
]
for r, note in enumerate(NOTES12, 16):
    ws12.cell(row=r, column=1, value=note)
    ws12.merge_cells(f"A{r}:N{r}")

for col in ws12.columns:
    ws12.column_dimensions[col[0].column_letter].width = 22

# ── Save Workbook 1 ───────────────────────────────────────────────────────────
PATH1 = "/Users/conor/Documents/case-study-data-analysis/case_study_tables.xlsx"
wb.save(PATH1)
print(f"Saved: {PATH1}")

# ══════════════════════════════════════════════════════════════════════════════
# WORKBOOK 2 — CHARTS
# ══════════════════════════════════════════════════════════════════════════════
dates_dt = [datetime.strptime(r[0], "%Y-%m-%d") for r in WEEKLY]
avg_ppms  = [r[1]  for r in WEEKLY]
run_hrs   = [r[3]  for r in WEEKLY]
plan_hrs  = [r[2]  for r in WEEKLY]
udt_hrs   = [r[4]  for r in WEEKLY]
completed = [r[7]  for r in WEEKLY]
passed    = [r[8]  for r in WEEKLY]

avails  = [rh/ph for rh, ph in zip(run_hrs, plan_hrs)]
perfs   = [ppm/CONTRACT_PPM for ppm in avg_ppms]
quals   = [p/c if c > 0 else 0 for p, c in zip(passed, completed)]
oees    = [a*p*q for a, p, q in zip(avails, perfs, quals)]

fig, axes = plt.subplots(2, 2, figsize=(16, 11))
fig.suptitle("Helix Unit 1 — Weekly Performance Trends (Aug–Nov 2025)",
             fontsize=15, fontweight='bold', y=0.98)
fig.patch.set_facecolor('#f8f8f8')

LABEL_FMT = mdates.DateFormatter('%d %b')
TICK_LOC  = mdates.WeekdayLocator(interval=2)

# ── Chart 1: Avg PPM ──────────────────────────────────────────────────────────
ax = axes[0, 0]
ax.plot(dates_dt, avg_ppms, color='#1565C0', linewidth=2.2, marker='o', markersize=5, label='Avg PPM')
ax.axhline(CONTRACT_PPM, color='#C62828', linestyle='--', linewidth=1.4, label='Contract target (300 PPM)')
ax.set_title("Avg Picks Per Minute (PPM)", fontweight='bold')
ax.set_ylabel("PPM")
ax.xaxis.set_major_formatter(LABEL_FMT)
ax.xaxis.set_major_locator(TICK_LOC)
plt.setp(ax.xaxis.get_majorticklabels(), rotation=40, ha='right', fontsize=8)
ax.legend(fontsize=8); ax.grid(True, alpha=0.3); ax.set_ylim(0, 360)
ax.set_facecolor('#fafafa')

# ── Chart 2: OEE & Components ────────────────────────────────────────────────
ax = axes[0, 1]
ax.plot(dates_dt, [v*100 for v in oees],   color='#2E7D32', linewidth=2.5, marker='o', markersize=5, label='OEE %')
ax.plot(dates_dt, [v*100 for v in avails], color='#1565C0', linewidth=1.5, marker='s', markersize=3, linestyle='--', alpha=0.75, label='Availability %')
ax.plot(dates_dt, [v*100 for v in perfs],  color='#E65100', linewidth=1.5, marker='^', markersize=3, linestyle='--', alpha=0.75, label='Performance %')
ax.plot(dates_dt, [v*100 for v in quals],  color='#6A1B9A', linewidth=1.5, marker='D', markersize=3, linestyle='--', alpha=0.75, label='Quality Yield %')
ax.axhline(50, color='#C62828', linestyle='--', linewidth=1.4, label='Contract OEE target (50%)')
ax.set_title("OEE and Components", fontweight='bold')
ax.set_ylabel("Percentage (%)")
ax.xaxis.set_major_formatter(LABEL_FMT)
ax.xaxis.set_major_locator(TICK_LOC)
plt.setp(ax.xaxis.get_majorticklabels(), rotation=40, ha='right', fontsize=8)
ax.legend(fontsize=7.5); ax.grid(True, alpha=0.3); ax.set_ylim(0, 115)
ax.set_facecolor('#fafafa')

# ── Chart 3: Quality Yield % ─────────────────────────────────────────────────
ax = axes[1, 0]
ax.fill_between(dates_dt, [v*100 for v in quals], alpha=0.15, color='#6A1B9A')
ax.plot(dates_dt, [v*100 for v in quals], color='#6A1B9A', linewidth=2.2, marker='o', markersize=5, label='Quality Yield %')
ax.axhline(80, color='#C62828', linestyle='--', linewidth=1.4, label='Contract quality target (80%)')
ax.set_title("Quality Yield % (Garments Passing QC / Completed)", fontweight='bold')
ax.set_ylabel("Quality Yield (%)")
ax.xaxis.set_major_formatter(LABEL_FMT)
ax.xaxis.set_major_locator(TICK_LOC)
plt.setp(ax.xaxis.get_majorticklabels(), rotation=40, ha='right', fontsize=8)
ax.legend(fontsize=8); ax.grid(True, alpha=0.3); ax.set_ylim(0, 105)
ax.set_facecolor('#fafafa')

# ── Chart 4: Unplanned Downtime Hours ────────────────────────────────────────
ax = axes[1, 1]
bar_labels = [d.strftime('%d %b') for d in dates_dt]
bars = ax.bar(bar_labels, udt_hrs, color='#EF9A9A', edgecolor='#C62828', linewidth=0.8)
# Colour bars: first 4 = worse period, last 4 = improving
for i, bar in enumerate(bars):
    bar.set_color('#EF9A9A' if i < 4 else ('#FFCC80' if i < 8 else '#A5D6A7'))
ax.set_title("Unplanned Downtime Hours per Week", fontweight='bold')
ax.set_ylabel("Hours")
plt.setp(ax.xaxis.get_majorticklabels(), rotation=40, ha='right', fontsize=8)
ax.grid(True, alpha=0.3, axis='y')
ax.set_facecolor('#fafafa')
# Add value labels
for bar, val in zip(bars, udt_hrs):
    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.2,
            f'{val}', ha='center', va='bottom', fontsize=7.5)

plt.tight_layout(rect=[0, 0, 1, 0.96])

# Save chart to buffer then embed in Excel
img_buf = io.BytesIO()
plt.savefig(img_buf, format='png', dpi=150, bbox_inches='tight', facecolor='#f8f8f8')
img_buf.seek(0)
plt.close()

wb2 = openpyxl.Workbook()
wb2.remove(wb2.active)

ws_chart = wb2.create_sheet("Performance Trends")
bold(ws_chart, 1, 1, "Helix Unit 1 — Weekly Performance Trends (Aug–Nov 2025)")
ws_chart.cell(row=2, column=1,
    value="OEE = Availability (Run Hrs / Planned Hrs) × Performance (Avg PPM / 300 PPM) × Quality Yield (Passed QC / Completed)")

xl_img = XLImage(img_buf)
xl_img.anchor = 'A4'
ws_chart.add_image(xl_img)

# Chart data sheet for reference
ws_data = wb2.create_sheet("Chart Data")
bold(ws_data, 1, 1, "Weekly performance data used for charts (all percentages rounded to 1 dp)")
header_row(ws_data, 2, ["Week Ending","Avg PPM","Planned Hrs","Run Hrs","Unplanned DT Hrs",
                         "Availability %","Performance %","Quality Yield %","OEE %"])
for r, (d, row, a, p, q, o) in enumerate(zip(
        [d.strftime('%Y-%m-%d') for d in dates_dt],
        WEEKLY, avails, perfs, quals, oees), 3):
    vals = [d, row[1], row[2], row[3], row[4],
            round(a*100,1), round(p*100,1), round(q*100,1), round(o*100,1)]
    for c, v in enumerate(vals, 1):
        ws_data.cell(row=r, column=c, value=v)
autowidth(ws_data, max_w=20)

PATH2 = "/Users/conor/Documents/case-study-data-analysis/case_study_charts.xlsx"
wb2.save(PATH2)
print(f"Saved: {PATH2}")
print("\nAll outputs complete.")
